#!/usr/bin/env python3
"""
odom_drive_log.py
─────────────────
Nhập số mét → xe đi đúng quãng đường đó → ghi odom vào Excel.

Timeline:
  [0.0s]              Phase WAIT_START: đứng yên 1s, ghi odom
  [1.0s]              Phase DRIVE: đi thẳng đến khi đủ khoảng cách, ghi odom
  [1.0s + travel_t]   Phase WAIT_END:  đứng yên 1s, ghi odom
  [xong]              Lưu Excel & thoát

Excel – sheet "Odom Log":
  time_s | phase | x_m | y_m | dist_traveled_m | vx_mps | wz_rads | target_m

Topics:
  Sub: /odom  (nav_msgs/Odometry)
  Pub: /drive (ackermann_msgs/AckermannDriveStamped)  ← topic F1TENTH chuẩn

Parameters:
  output_path  : ~/odom_drive_log.xlsx
  drive_topic  : /drive
  speed_mps    : 0.5   (tốc độ đi, m/s)
  log_hz       : 20    (tần số ghi dữ liệu)
  wait_sec     : 1.0   (thời gian đứng yên đầu/cuối)

Yêu cầu: pip install openpyxl
"""

import math
import os
import sys
import threading
import time

import rclpy
from rclpy.node import Node
from rclpy.qos import QoSProfile, QoSReliabilityPolicy, QoSHistoryPolicy
from nav_msgs.msg import Odometry

try:
    from ackermann_msgs.msg import AckermannDriveStamped
    _HAS_ACKERMANN = True
except ImportError:
    _HAS_ACKERMANN = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("pip install openpyxl")


# ── style ─────────────────────────────────────────────────────────────────────
_FILL_BLUE  = PatternFill("solid", fgColor="1F4E79")
_FILL_AMBER = PatternFill("solid", fgColor="7B3F00")
_FILL_GREEN = PatternFill("solid", fgColor="1E5631")
_FONT_HDR   = Font(bold=True, color="FFFFFF")
_CENTER     = Alignment(horizontal="center")

_PHASE_COLOR = {
    "WAIT_START": _FILL_BLUE,
    "DRIVE":      _FILL_AMBER,
    "WAIT_END":   _FILL_GREEN,
}

# Phase constants
IDLE       = "IDLE"
WAIT_START = "WAIT_START"
DRIVE      = "DRIVE"
WAIT_END   = "WAIT_END"
DONE       = "DONE"


class OdomDriveLog(Node):
    def __init__(self):
        super().__init__("odom_drive_log_node")

        # ── params ────────────────────────────────────────────────────────
        def P(n, v):
            self.declare_parameter(n, v)
            return self.get_parameter(n).value

        self.output_path = P("output_path", os.path.expanduser("~/odom_drive_log.xlsx"))
        self.drive_topic = P("drive_topic", "/drive")
        self.speed_mps   = float(P("speed_mps",  0.5))
        self.log_hz      = float(P("log_hz",      20.0))
        self.wait_sec    = float(P("wait_sec",    1.0))

        if not _HAS_ACKERMANN:
            self.get_logger().error(
                "Thiếu ackermann_msgs! Cài: sudo apt install ros-humble-ackermann-msgs"
            )
            sys.exit(1)

        self.get_logger().info(f"Output : {self.output_path}")
        self.get_logger().info(f"Topic  : {self.drive_topic}")
        self.get_logger().info(f"Speed  : {self.speed_mps} m/s")

        # ── state machine ────────────────────────────────────────────────
        self.phase       = IDLE
        self.target_m    = 0.0
        self.phase_start = 0.0        # time.monotonic()
        self.x0 = self.y0 = None     # vị trí bắt đầu phase DRIVE
        self.start_time  = 0.0       # time.monotonic() khi bắt đầu toàn bộ

        # ── odom latest ──────────────────────────────────────────────────
        self.odom_lock = threading.Lock()
        self.latest_odom: Odometry = None

        # ── workbook ─────────────────────────────────────────────────────
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Odom Log"
        self._init_header()
        self.row_count = 0

        # ── pub / sub ────────────────────────────────────────────────────
        self.pub_drive = self.create_publisher(AckermannDriveStamped, self.drive_topic, 10)

        # /odom thường publish BEST_EFFORT → dùng cùng QoS để match
        _qos_best = QoSProfile(
            reliability=QoSReliabilityPolicy.BEST_EFFORT,
            history=QoSHistoryPolicy.KEEP_LAST,
            depth=10,
        )
        self.create_subscription(Odometry, "/odom", self._odom_cb, _qos_best)

        # ── log timer (khởi động sau khi có lệnh) ────────────────────────
        self._log_timer = None

        # ── keyboard thread ───────────────────────────────────────────────
        threading.Thread(target=self._kb_loop, daemon=True).start()

        print("\n" + "="*60)
        print("  ODOM DRIVE LOG")
        print("="*60)
        print(f"  Nhấn Enter → nhập số mét → xe đi → lưu Excel")
        print(f"  Tốc độ: {self.speed_mps} m/s | Wait: {self.wait_sec}s đầu/cuối")
        print("="*60 + "\n")

    # ── odom callback ────────────────────────────────────────────────────────
    def _odom_cb(self, msg: Odometry):
        with self.odom_lock:
            self.latest_odom = msg

    # ── keyboard thread ──────────────────────────────────────────────────────
    def _kb_loop(self):
        while rclpy.ok():
            try:
                input("  → Nhấn Enter để bắt đầu ... ")
            except EOFError:
                break

            if self.phase not in (IDLE, DONE):
                print("  ⚠  Đang chạy, chờ xong rồi nhập tiếp.\n")
                continue

            raw = input("  Khoảng cách cần đi (m): ").strip()
            try:
                dist = float(raw)
                if dist <= 0:
                    raise ValueError
            except ValueError:
                print("  ⚠  Nhập số dương, vd: 0.3\n")
                continue

            self._start_run(dist)

    # ── start a run ──────────────────────────────────────────────────────────
    def _start_run(self, target_m: float):
        with self.odom_lock:
            if self.latest_odom is None:
                print("  ⚠  Chưa nhận được /odom! Kiểm tra kết nối.\n")
                return

        self.target_m    = target_m
        self.start_time  = time.monotonic()
        self.phase_start = self.start_time
        self.phase       = WAIT_START
        self.x0 = self.y0 = None

        print(f"\n  ▶ Bắt đầu: target={target_m} m")
        print(f"  [WAIT_START] Đứng yên {self.wait_sec}s ...")

        # Tạo timer log nếu chưa có
        if self._log_timer is None:
            period = 1.0 / self.log_hz
            self._log_timer = self.create_timer(period, self._log_tick)

    # ── log timer callback ───────────────────────────────────────────────────
    def _log_tick(self):
        if self.phase in (IDLE, DONE):
            return

        now = time.monotonic()
        elapsed_total = round(now - self.start_time, 4)
        phase_elapsed = now - self.phase_start

        with self.odom_lock:
            odom = self.latest_odom
        if odom is None:
            return

        px = odom.pose.pose.position.x
        py = odom.pose.pose.position.y
        vx = odom.twist.twist.linear.x
        wz = odom.twist.twist.angular.z

        # khoảng cách đã đi kể từ khi bắt đầu phase DRIVE
        if self.x0 is not None:
            dist_traveled = math.sqrt((px - self.x0)**2 + (py - self.y0)**2)
        else:
            dist_traveled = 0.0

        # ── state transitions ────────────────────────────────────────────
        if self.phase == WAIT_START:
            # Gửi lệnh dừng đều đặn để giữ xe đứng yên
            self._publish_drive(0.0)
            if phase_elapsed >= self.wait_sec:
                self.phase       = DRIVE
                self.phase_start = now
                self.x0, self.y0 = px, py
                print(f"  [DRIVE] Đang đi {self.target_m} m ...")

        elif self.phase == DRIVE:
            remaining = self.target_m - dist_traveled

            if remaining <= 0:
                # Đã đến đích → dừng
                self._publish_drive(0.0)
                self.phase       = WAIT_END
                self.phase_start = now
                print(f"  [WAIT_END] Đã đi {dist_traveled:.4f} m, đứng yên {self.wait_sec}s ...")
            else:
                # ── Vùng giảm tốc: 30% cuối quãng đường ──────────────
                decel_start = self.target_m * 0.30   # bắt đầu giảm tốc
                min_speed   = 0.15                    # tốc độ tối thiểu (m/s)

                if remaining < decel_start:
                    # Tỷ lệ 0→1 khi remaining đi từ decel_start→0
                    ratio = remaining / decel_start
                    cmd_speed = max(min_speed, self.speed_mps * ratio)
                else:
                    cmd_speed = self.speed_mps

                self._publish_drive(cmd_speed)

        elif self.phase == WAIT_END:
            self._publish_drive(0.0)
            if phase_elapsed >= self.wait_sec:
                self.phase = DONE
                print(f"\n  ✓ Hoàn thành! dist_traveled={dist_traveled:.4f} m")
                self._save()
                return

        # ── ghi 1 hàng Excel ─────────────────────────────────────────────
        self.ws.append([
            elapsed_total,
            self.phase,
            round(px, 6),
            round(py, 6),
            round(dist_traveled, 6),
            round(vx, 6),
            round(wz, 6),
            self.target_m,
        ])
        self.row_count += 1

    # ── publish drive command ────────────────────────────────────────────────
    def _publish_drive(self, speed: float):
        msg = AckermannDriveStamped()
        msg.header.stamp = self.get_clock().now().to_msg()
        msg.drive.speed              = float(speed)
        msg.drive.steering_angle     = 0.0
        self.pub_drive.publish(msg)

    # ── header ───────────────────────────────────────────────────────────────
    def _init_header(self):
        cols = ["time_s", "phase",
                "x_m", "y_m", "dist_traveled_m",
                "vx_mps", "wz_rads", "target_m"]
        self.ws.append(cols)
        for cell in self.ws[1]:
            cell.font = _FONT_HDR
            cell.fill = _FILL_BLUE
            cell.alignment = _CENTER
        widths = [12, 14, 14, 14, 18, 12, 12, 12]
        for i, w in enumerate(widths, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = w
        self.ws.freeze_panes = "A2"

    # ── save ─────────────────────────────────────────────────────────────────
    def _save(self):
        # Tô màu cột phase theo từng giai đoạn
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row):
            phase_val = row[1].value
            fill = _PHASE_COLOR.get(phase_val)
            if fill:
                for cell in row:
                    cell.fill = fill

        try:
            self.wb.save(self.output_path)
            print(f"  ✓ Excel ({self.row_count} hàng): {self.output_path}\n")
        except Exception as e:
            self.get_logger().error(f"Lưu Excel: {e}")


# ── main ─────────────────────────────────────────────────────────────────────
def main(args=None):
    rclpy.init(args=args)
    node = OdomDriveLog()
    try:
        rclpy.spin(node)
    except KeyboardInterrupt:
        pass
    finally:
        # Dừng xe trước khi tắt (kiểm tra context còn hợp lệ)
        try:
            if rclpy.ok():
                node._publish_drive(0.0)
        except Exception:
            pass
        if node.row_count > 0 and node.phase != DONE:
            node._save()
        node.destroy_node()
        if rclpy.ok():
            rclpy.shutdown()


if __name__ == "__main__":
    main()
