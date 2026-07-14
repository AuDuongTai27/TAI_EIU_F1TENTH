#!/usr/bin/env python3
"""
depth_lidar_single.py  –  Single-shot mode
──────────────────────────────────────────
Dùng để so sánh số liệu thật với số liệu cảm biến đo được.

Quy trình:
  1. Đặt vật ở khoảng cách cần đo (1m / 5m / 10m ...)
  2. Nhấn Enter → nhập khoảng cách thật (vd: 1.0)
  3. Node chụp 1 bức: color + depth + LiDAR → ghi 1 hàng Excel
  4. Lặp lại cho các mốc khác → Ctrl+C để lưu

Excel – sheet "Measurements":
  shot | distance_real_m | timestamp_sec |
  depth_center_m | depth_roi_mean_m | depth_roi_std_m |
  lidar_mean_m | lidar_min_m | lidar_count |
  color_file | depth_file

Yêu cầu: pip install openpyxl numpy
"""

import math, os, struct, threading, zlib
import rclpy
from rclpy.node import Node
from sensor_msgs.msg import Image, LaserScan

try:
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise ImportError(f"pip install openpyxl numpy  ({e})")

# ── style ─────────────────────────────────────────────────────────────────────
_FILL_BLUE = PatternFill("solid", fgColor="1F4E79")
_FONT_HDR  = Font(bold=True, color="FFFFFF")
_CENTER    = Alignment(horizontal="center")

def _style_header(ws, fill):
    for c in ws[1]:
        c.font = _FONT_HDR; c.fill = fill; c.alignment = _CENTER

# ── image utils (pure Python, không cần cv2) ──────────────────────────────────
def _decode_depth(msg: Image) -> np.ndarray:
    if msg.encoding == "32FC1":
        return np.frombuffer(bytes(msg.data), np.float32).reshape(msg.height, msg.width).copy()
    elif msg.encoding in ("16UC1", "mono16"):
        return (np.frombuffer(bytes(msg.data), np.uint16)
                  .reshape(msg.height, msg.width).astype(np.float32) / 1000.0)
    raise RuntimeError(f"Encoding không hỗ trợ: {msg.encoding}")

def _to_png(arr_rgb: np.ndarray) -> bytes:
    H, W = arr_rgb.shape[:2]
    def chunk(t, d):
        c = t + d
        return struct.pack(">I", len(d)) + c + struct.pack(">I", zlib.crc32(c) & 0xFFFFFFFF)
    ihdr = chunk(b"IHDR", struct.pack(">IIBBBBB", W, H, 8, 2, 0, 0, 0))
    raw  = b"".join(b"\x00" + arr_rgb[r].tobytes() for r in range(H))
    idat = chunk(b"IDAT", zlib.compress(raw, 6))
    iend = chunk(b"IEND", b"")
    return b"\x89PNG\r\n\x1a\n" + ihdr + idat + iend

def _color_to_png(msg: Image) -> bytes:
    arr = np.frombuffer(bytes(msg.data), np.uint8).reshape(msg.height, msg.width, 3)
    if msg.encoding == "bgr8":
        arr = arr[:, :, ::-1].copy()
    return _to_png(arr)

def _depth_to_png(arr: np.ndarray, d_max: float = 10.0) -> bytes:
    valid = np.isfinite(arr) & (arr > 0)
    g = np.zeros_like(arr, dtype=np.uint8)
    g[valid] = np.clip(255 * (1.0 - arr[valid] / d_max), 0, 255).astype(np.uint8)
    return _to_png(np.stack([g, g, g], axis=-1))


class DepthLidarSingle(Node):
    def __init__(self):
        super().__init__("depth_lidar_single_node")

        # ── params ────────────────────────────────────────────────────────
        def P(n, v): self.declare_parameter(n, v); return self.get_parameter(n).value

        self.output_path     = P("output_path", os.path.expanduser("~/depth_lidar_single.xlsx"))
        self.photo_dir       = P("photo_dir",   os.path.expanduser("~/depth_lidar_photos"))
        self.depth_topic     = P("depth_topic", "/camera/camera/aligned_depth_to_color/image_raw")
        self.color_topic     = P("color_topic", "/camera/camera/color/image_raw")
        self.roi_radius      = int(P("roi_radius_px",   20))
        self.front_angle_deg = float(P("front_angle_deg", 10.0))
        self.depth_vis_max   = float(P("depth_vis_max_m", 10.0))

        os.makedirs(self.photo_dir, exist_ok=True)

        self.get_logger().info(f"Output  : {self.output_path}")
        self.get_logger().info(f"Photos  : {self.photo_dir}")
        self.get_logger().info(f"ROI     : ±{self.roi_radius}px | LiDAR front: ±{self.front_angle_deg}°")

        # ── state ─────────────────────────────────────────────────────────
        self.latest_depth: np.ndarray = None
        self.latest_color: Image      = None
        self.latest_scan:  LaserScan  = None
        self._dlck = threading.Lock()
        self._clck = threading.Lock()
        self._slck = threading.Lock()
        self.shot_count = 0

        # ── workbook ──────────────────────────────────────────────────────
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Measurements"
        self._init_header()

        # ── subscribers ───────────────────────────────────────────────────
        self.create_subscription(Image,     self.depth_topic, self._depth_cb, 10)
        self.create_subscription(Image,     self.color_topic, self._color_cb, 10)
        self.create_subscription(LaserScan, "/scan",          self._scan_cb,  10)

        # ── keyboard thread ───────────────────────────────────────────────
        threading.Thread(target=self._kb_loop, daemon=True).start()

        print("\n" + "="*60)
        print("  SINGLE-SHOT: DEPTH + COLOR + LIDAR")
        print("="*60)
        print("  Đặt vật ở khoảng cách cần đo,")
        print("  nhấn  [Enter]  → nhập khoảng cách thật (m)")
        print("  → node chụp 1 bức + ghi Excel")
        print("  Nhấn  [Ctrl+C]  để lưu và thoát")
        print("="*60 + "\n")

    # ── callbacks ────────────────────────────────────────────────────────────
    def _depth_cb(self, msg: Image):
        try:
            arr = _decode_depth(msg)
            with self._dlck: self.latest_depth = arr
        except Exception as e:
            self.get_logger().error(f"Depth: {e}")

    def _color_cb(self, msg: Image):
        with self._clck: self.latest_color = msg

    def _scan_cb(self, msg: LaserScan):
        with self._slck: self.latest_scan = msg

    # ── keyboard ─────────────────────────────────────────────────────────────
    def _kb_loop(self):
        while rclpy.ok():
            try:
                input("  → Nhấn Enter để chụp ... ")
            except EOFError:
                break

            # snapshot
            with self._dlck: depth = self.latest_depth.copy() if self.latest_depth is not None else None
            with self._clck: color = self.latest_color
            with self._slck: scan  = self.latest_scan

            missing = ([" Depth"] if depth is None else []) + ([" LiDAR"] if scan is None else [])
            if missing:
                print(f"  ⚠  Chưa có dữ liệu:{', '.join(missing)}. Thử lại.\n")
                continue

            # hỏi khoảng cách thật
            raw = input("  Khoảng cách thật (m): ").strip()
            try:
                dist_real = float(raw)
            except ValueError:
                print("  ⚠  Nhập số thực, vd: 1.0  →  bỏ qua.\n")
                continue

            self._capture(depth, color, scan, dist_real)

    # ── capture 1 shot ───────────────────────────────────────────────────────
    def _capture(self, depth_arr, color_msg, scan_msg, dist_real: float):
        self.shot_count += 1
        shot = self.shot_count

        # ── depth ROI ──────────────────────────────────────────────────
        H, W = depth_arr.shape[:2]
        cx, cy, r = W//2, H//2, self.roi_radius
        roi  = depth_arr[max(0,cy-r):cy+r+1, max(0,cx-r):cx+r+1]
        mask = np.isfinite(roi) & (roi > 0.01)
        vals = roi[mask]
        if vals.size:
            cv       = roi[min(r, cy), min(r, cx)]
            d_center = round(float(cv), 4) if (np.isfinite(cv) and cv > 0) else None
            d_mean   = round(float(np.mean(vals)), 4)
            d_std    = round(float(np.std(vals)),  4)
        else:
            d_center = d_mean = d_std = None

        # ── LiDAR front ────────────────────────────────────────────────
        fa = math.radians(self.front_angle_deg)
        lr = [v for i, v in enumerate(scan_msg.ranges)
              if math.isfinite(v) and v > 0
              and abs(scan_msg.angle_min + i * scan_msg.angle_increment) <= fa]
        if lr:
            l_mean = round(float(np.mean(lr)), 4)
            l_min  = round(float(min(lr)),     4)
            l_n    = len(lr)
        else:
            l_mean = l_min = l_n = None

        # ── timestamp ──────────────────────────────────────────────────
        st = scan_msg.header.stamp
        ts = round(st.sec + st.nanosec * 1e-9, 6)

        # ── lưu ảnh ────────────────────────────────────────────────────
        cfn = f"shot_{shot:03d}_color.png"
        dfn = f"shot_{shot:03d}_depth.png"
        cpath = os.path.join(self.photo_dir, cfn)
        dpath = os.path.join(self.photo_dir, dfn)

        if color_msg is not None:
            try:
                with open(cpath, "wb") as f: f.write(_color_to_png(color_msg))
            except Exception as e:
                self.get_logger().error(f"Lưu color: {e}"); cfn = "ERROR"
        else:
            cfn = "—"

        try:
            with open(dpath, "wb") as f:
                f.write(_depth_to_png(depth_arr, d_max=self.depth_vis_max))
        except Exception as e:
            self.get_logger().error(f"Lưu depth: {e}"); dfn = "ERROR"

        # ── ghi Excel ──────────────────────────────────────────────────
        self.ws.append([shot, dist_real, ts,
                        d_center, d_mean, d_std,
                        l_mean, l_min, l_n, cfn, dfn])

        # ── in kết quả ─────────────────────────────────────────────────
        err_d = round(d_mean - dist_real, 4) if d_mean is not None else None
        err_l = round(l_mean - dist_real, 4) if l_mean is not None else None
        print(f"\n  ✓ Shot #{shot}  |  thật = {dist_real} m")
        print(f"    Depth  : center={d_center} m | mean={d_mean} m | std={d_std} m | lệch={err_d} m")
        print(f"    LiDAR  : mean={l_mean} m | min={l_min} m | lệch={err_l} m")
        print(f"    Ảnh    : {cfn}  |  {dfn}\n")

    # ── header ───────────────────────────────────────────────────────────────
    def _init_header(self):
        cols = ["shot", "distance_real_m", "timestamp_sec",
                "depth_center_m", "depth_roi_mean_m", "depth_roi_std_m",
                "lidar_mean_m", "lidar_min_m", "lidar_count",
                "color_file", "depth_file"]
        self.ws.append(cols)
        _style_header(self.ws, _FILL_BLUE)
        for i, w in enumerate([8, 18, 18, 16, 18, 16, 14, 12, 12, 28, 28], 1):
            self.ws.column_dimensions[get_column_letter(i)].width = w
        self.ws.freeze_panes = "A2"

    # ── save ─────────────────────────────────────────────────────────────────
    def save(self):
        try:
            self.wb.save(self.output_path)
            print(f"\n  ✓ Excel ({self.shot_count} hàng): {self.output_path}")
            print(f"  ✓ Ảnh: {self.photo_dir}/")
        except Exception as e:
            self.get_logger().error(f"Lưu Excel: {e}")


# ── main ─────────────────────────────────────────────────────────────────────
def main(args=None):
    rclpy.init(args=args)
    node = DepthLidarSingle()
    try:
        rclpy.spin(node)
    except KeyboardInterrupt:
        pass
    finally:
        if node.shot_count > 0:
            node.save()
        else:
            print("  ⚠  Chưa có lần đo nào.")
        node.destroy_node()
        if rclpy.ok():
            rclpy.shutdown()

if __name__ == "__main__":
    main()
