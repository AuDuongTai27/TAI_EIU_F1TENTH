#!/usr/bin/env python3
"""
servo_steer_log.py
──────────────────
Kiểm tra servo bằng cách lần lượt:
  1. Rẽ TRÁI  (steering = +max_angle)  trong phase_sec giây
  2. Đi THẲNG (steering = 0.0)         trong phase_sec giây
  3. Rẽ PHẢI  (steering = -max_angle)  trong phase_sec giây

Trong suốt quá trình ghi dữ liệu steering vào Excel.

Excel – sheet "Servo Log":
  time_s | phase | cmd_angle_rad | cmd_angle_deg |
  actual_angle_rad | actual_angle_deg | error_deg

Topics:
  Pub: /drive                          (ackermann_msgs/AckermannDriveStamped)
  Sub: /sensors/servo_position_command (std_msgs/Float64)   ← feedback thực tế

Parameters:
  output_path   : ~/servo_steer_log.xlsx
  drive_topic   : /drive
  max_angle_deg : 20.0   (góc lái tối đa, độ)
  phase_sec     : 5.0    (thời gian mỗi phase, giây)
  drive_speed   : 0.0    (tốc độ xe khi test, m/s – mặc định đứng yên)
  log_hz        : 20.0   (tần số ghi log, Hz)

Yêu cầu: pip install openpyxl
"""

import math
import os
import threading
import time

import rclpy
from rclpy.node import Node
from rclpy.qos import QoSProfile, QoSReliabilityPolicy, QoSHistoryPolicy
from std_msgs.msg import Float64

try:
    from ackermann_msgs.msg import AckermannDriveStamped
    _HAS_ACKERMANN = True
except ImportError:
    _HAS_ACKERMANN = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("pip install openpyxl")


# ── Phase constants ───────────────────────────────────────────────────────────
IDLE     = "IDLE"
LEFT     = "LEFT"
STRAIGHT = "STRAIGHT"
RIGHT    = "RIGHT"
DONE     = "DONE"

SEQUENCE = [LEFT, STRAIGHT, RIGHT]

# ── Colors ────────────────────────────────────────────────────────────────────
_FILL_HDR      = PatternFill("solid", fgColor="1F4E79")
_FILL_LEFT     = PatternFill("solid", fgColor="1A3A6B")   # xanh đậm
_FILL_STRAIGHT = PatternFill("solid", fgColor="1E5631")   # xanh lá
_FILL_RIGHT    = PatternFill("solid", fgColor="7B1A1A")   # đỏ đậm
_FONT_HDR      = Font(bold=True, color="FFFFFF")
_CENTER        = Alignment(horizontal="center")

_PHASE_FILL = {LEFT: _FILL_LEFT, STRAIGHT: _FILL_STRAIGHT, RIGHT: _FILL_RIGHT}
_PHASE_ANGLE = {LEFT: 1, STRAIGHT: 0, RIGHT: -1}   # multiplier × max_angle


class ServoSteerLog(Node):
    def __init__(self):
        super().__init__("servo_steer_log_node")

        # ── params ────────────────────────────────────────────────────────
        def P(n, v):
            self.declare_parameter(n, v)
            return self.get_parameter(n).value

        self.output_path  = P("output_path",   os.path.expanduser("~/servo_steer_log.xlsx"))
        self.drive_topic  = P("drive_topic",   "/drive")
        self.max_angle_r  = math.radians(float(P("max_angle_deg", 20.0)))
        self.phase_sec    = float(P("phase_sec",    5.0))
        self.drive_speed  = float(P("drive_speed",  0.0))
        self.log_hz       = float(P("log_hz",       20.0))

        if not _HAS_ACKERMANN:
            self.get_logger().error("Thiếu ackermann_msgs! sudo apt install ros-humble-ackermann-msgs")
            raise SystemExit(1)

        self.get_logger().info(f"Output     : {self.output_path}")
        self.get_logger().info(f"max_angle  : {math.degrees(self.max_angle_r):.1f}°")
        self.get_logger().info(f"phase_sec  : {self.phase_sec}s × 3 phases")

        # ── state ─────────────────────────────────────────────────────────
        self.phase        = IDLE
        self.phase_idx    = -1            # index trong SEQUENCE
        self.phase_start  = 0.0
        self.run_start    = 0.0
        self.actual_angle = 0.0           # feedback từ servo
        self._angle_lock  = threading.Lock()

        # ── workbook ──────────────────────────────────────────────────────
        self.wb       = Workbook()
        self.ws       = self.wb.active
        self.ws.title = "Servo Log"
        self._init_header()
        self.row_count = 0

        # ── pub / sub ─────────────────────────────────────────────────────
        self.pub_drive = self.create_publisher(AckermannDriveStamped, self.drive_topic, 10)

        _qos = QoSProfile(
            reliability=QoSReliabilityPolicy.BEST_EFFORT,
            history=QoSHistoryPolicy.KEEP_LAST,
            depth=10,
        )
        self.create_subscription(Float64, "/sensors/servo_position_command",
                                 self._servo_cb, _qos)

        # ── keyboard thread ───────────────────────────────────────────────
        threading.Thread(target=self._kb_loop, daemon=True).start()

        print("\n" + "="*62)
        print("  SERVO STEER LOG")
        print("="*62)
        print(f"  Trình tự: TRÁI {self.phase_sec}s → THẲNG {self.phase_sec}s → PHẢI {self.phase_sec}s")
        print(f"  Góc tối đa: ±{math.degrees(self.max_angle_r):.1f}°")
        print(f"  Tốc độ xe : {self.drive_speed} m/s")
        print("="*62 + "\n")

    # ── servo feedback ───────────────────────────────────────────────────────
    def _servo_cb(self, msg: Float64):
        with self._angle_lock:
            # VESC servo position: 0.0–1.0 → map sang rad
            # 0.5 = thẳng, 0.0 = max right, 1.0 = max left (thông thường)
            # Hoặc topic đã là radian – tùy firmware
            # Giữ raw value, ghi cả 2 vào Excel để người dùng tự map
            self.actual_angle = msg.data

    # ── keyboard ─────────────────────────────────────────────────────────────
    def _kb_loop(self):
        while rclpy.ok():
            try:
                input("  → Nhấn Enter để bắt đầu kiểm tra servo ... ")
            except EOFError:
                break

            if self.phase not in (IDLE, DONE):
                print("  ⚠  Đang chạy, chờ xong.\n")
                continue

            self._start_sequence()

    # ── start ─────────────────────────────────────────────────────────────────
    def _start_sequence(self):
        self.phase_idx   = 0
        self.phase       = SEQUENCE[0]
        self.phase_start = time.monotonic()
        self.run_start   = self.phase_start

        print(f"\n  ▶ Bắt đầu  [LEFT] {self.phase_sec}s ...")
        self._log_timer = self.create_timer(1.0 / self.log_hz, self._log_tick)

    # ── timer tick ────────────────────────────────────────────────────────────
    def _log_tick(self):
        if self.phase in (IDLE, DONE):
            return

        now           = time.monotonic()
        elapsed_total = round(now - self.run_start,   4)
        phase_elapsed = now - self.phase_start

        # ── góc lệnh theo phase ──────────────────────────────────────────
        cmd_rad = self.max_angle_r * _PHASE_ANGLE[self.phase]
        cmd_deg = round(math.degrees(cmd_rad), 3)

        # ── publish drive ────────────────────────────────────────────────
        self._send_drive(cmd_rad)

        # ── đọc VESC servo feedback (0–1, không phải encoder vật lý) ────
        with self._angle_lock:
            servo_raw = self.actual_angle   # 0=full right, 0.5=thẳng, 1=full left

        # ── ghi Excel ────────────────────────────────────────────────────
        self.ws.append([
            elapsed_total,
            self.phase,
            round(cmd_rad, 6),
            cmd_deg,
            round(servo_raw, 6),
        ])
        self.row_count += 1

        # ── chuyển phase ─────────────────────────────────────────────────
        if phase_elapsed >= self.phase_sec:
            self.phase_idx += 1
            if self.phase_idx >= len(SEQUENCE):
                # Xong hết
                self._send_drive(0.0)     # trả về thẳng
                self.phase = DONE
                print(f"\n  ✓ Hoàn thành 3 phase!")
                self._save()
                self._log_timer.cancel()
            else:
                self.phase       = SEQUENCE[self.phase_idx]
                self.phase_start = now
                label = {"LEFT": "TRÁI", "STRAIGHT": "THẲNG", "RIGHT": "PHẢI"}
                print(f"  ▶ [{self.phase}] {label[self.phase]} {self.phase_sec}s ...")

    # ── publish drive ─────────────────────────────────────────────────────────
    def _send_drive(self, steering_rad: float):
        msg = AckermannDriveStamped()
        msg.header.stamp         = self.get_clock().now().to_msg()
        msg.drive.speed          = float(self.drive_speed)
        msg.drive.steering_angle = float(steering_rad)
        self.pub_drive.publish(msg)

    # ── header ────────────────────────────────────────────────────────────────
    def _init_header(self):
        cols = [
            "time_s", "phase",
            "cmd_angle_rad", "cmd_angle_deg",
            "servo_raw_vesc",   # VESC echo: 0=full right, 0.5=thẳng, 1=full left (không phải encoder)
        ]
        self.ws.append(cols)
        for cell in self.ws[1]:
            cell.font = _FONT_HDR; cell.fill = _FILL_HDR; cell.alignment = _CENTER
        widths = [12, 12, 16, 16, 18]
        for i, w in enumerate(widths, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = w
        self.ws.freeze_panes = "A2"

    # ── save ──────────────────────────────────────────────────────────────────
    def _save(self):
        # Tô màu theo phase
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row):
            fill = _PHASE_FILL.get(row[1].value)
            if fill:
                for cell in row:
                    cell.fill = fill
        try:
            self.wb.save(self.output_path)
            print(f"  ✓ Excel ({self.row_count} hàng): {self.output_path}")
        except Exception as e:
            self.get_logger().error(f"Lưu Excel: {e}")


# ── main ──────────────────────────────────────────────────────────────────────
def main(args=None):
    rclpy.init(args=args)
    node = ServoSteerLog()
    try:
        rclpy.spin(node)
    except KeyboardInterrupt:
        pass
    finally:
        try:
            if rclpy.ok():
                node._send_drive(0.0)    # trả về thẳng khi tắt
        except Exception:
            pass
        if node.row_count > 0 and node.phase != DONE:
            node._save()
        node.destroy_node()
        if rclpy.ok():
            rclpy.shutdown()


if __name__ == "__main__":
    main()
