#!/usr/bin/env python3
"""
Node subscriber nhận dữ liệu LiDAR từ topic /scan và ghi vào file Excel (.xlsx).

Cấu trúc file Excel gồm 2 sheet:
  ┌─ Sheet 1 "Scan Summary" ─────────────────────────────────────────────────┐
  │  1 hàng = 1 scan message                                                  │
  │  Cột: msg_index | timestamp | angle_min/max | range_min/max |             │
  │        num_ranges | stat_min | stat_max | stat_mean | stat_std |          │
  │        closest_range | closest_angle_deg                                  │
  └───────────────────────────────────────────────────────────────────────────┘
  ┌─ Sheet 2 "Ranges" (dạng dọc / tidy) ────────────────────────────────────┐
  │  1 hàng = 1 beam của 1 scan                                               │
  │  Cột: msg_index | timestamp_sec | beam_index | angle_deg | range_m        │
  │  → dễ lọc, sort, pivot, vẽ chart trong Excel                             │
  └───────────────────────────────────────────────────────────────────────────┘

Yêu cầu: pip install openpyxl
"""

import math
import os

import rclpy
from rclpy.node import Node
from sensor_msgs.msg import LaserScan

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Thiếu openpyxl. Cài bằng: pip install openpyxl")


# ─── Màu header ───────────────────────────────────────────────────────────────
_HEADER_FILL_BLUE  = PatternFill("solid", fgColor="1F4E79")   # Sheet 1
_HEADER_FILL_GREEN = PatternFill("solid", fgColor="1E5631")   # Sheet 2
_HEADER_FONT       = Font(bold=True, color="FFFFFF")
_CENTER            = Alignment(horizontal="center")


class LidarToExcel(Node):
    def __init__(self):
        super().__init__("lidar_to_excel_node")
        self.get_logger().info("lidar_to_excel_node khởi động.")

        # ── Parameters ────────────────────────────────────────────────────
        self.declare_parameter(
            "output_path",
            os.path.join(os.path.expanduser("~"), "lidar_data.xlsx"),
        )
        self.output_path: str = (
            self.get_parameter("output_path").get_parameter_value().string_value
        )

        # 0 = không giới hạn
        self.declare_parameter("max_messages", 0)
        self.max_messages: int = (
            self.get_parameter("max_messages").get_parameter_value().integer_value
        )

        self.get_logger().info(f"Output: {self.output_path}")

        # ── Workbook ──────────────────────────────────────────────────────
        self.wb = Workbook()

        # Sheet 1 – Scan Summary
        self.ws_summary = self.wb.active
        self.ws_summary.title = "Scan Summary"
        self._init_summary_header()

        # Sheet 2 – Ranges (tidy / dạng dọc)
        self.ws_ranges = self.wb.create_sheet("Ranges")
        self._init_ranges_header()

        self.message_count = 0

        # ── Subscriber ────────────────────────────────────────────────────
        self.subscription = self.create_subscription(
            LaserScan, "scan", self.scan_callback, 10
        )

    # ─── Header helpers ───────────────────────────────────────────────────────
    def _style_header_row(self, sheet, fill):
        for cell in sheet[1]:
            cell.font      = _HEADER_FONT
            cell.fill      = fill
            cell.alignment = _CENTER

    def _init_summary_header(self):
        cols = [
            "msg_index",
            "timestamp_sec",
            "angle_min_deg",
            "angle_max_deg",
            "angle_increment_deg",
            "range_min_m",
            "range_max_m",
            "num_beams",
            "stat_min_m",
            "stat_max_m",
            "stat_mean_m",
            "stat_std_m",
            "closest_range_m",
            "closest_angle_deg",
        ]
        self.ws_summary.append(cols)
        self._style_header_row(self.ws_summary, _HEADER_FILL_BLUE)
        # Đặt độ rộng cột cố định
        for i, _ in enumerate(cols, 1):
            self.ws_summary.column_dimensions[get_column_letter(i)].width = 20

    def _init_ranges_header(self):
        cols = ["msg_index", "timestamp_sec", "beam_index", "angle_deg", "range_m"]
        self.ws_ranges.append(cols)
        self._style_header_row(self.ws_ranges, _HEADER_FILL_GREEN)
        widths = [12, 18, 12, 14, 12]
        for i, w in enumerate(widths, 1):
            self.ws_ranges.column_dimensions[get_column_letter(i)].width = w

    # ─── Callback ─────────────────────────────────────────────────────────────
    def scan_callback(self, msg: LaserScan):
        self.message_count += 1
        idx = self.message_count

        stamp = msg.header.stamp
        ts    = round(stamp.sec + stamp.nanosec * 1e-9, 6)

        ranges = msg.ranges
        n      = len(ranges)

        # Lọc bỏ inf / nan để tính thống kê
        valid = [r for r in ranges if math.isfinite(r) and r > 0]

        if valid:
            stat_min  = round(min(valid), 4)
            stat_max  = round(max(valid), 4)
            stat_mean = round(sum(valid) / len(valid), 4)
            variance  = sum((r - stat_mean) ** 2 for r in valid) / len(valid)
            stat_std  = round(math.sqrt(variance), 4)

            # Beam gần nhất
            min_idx         = ranges.index(min(valid))
            closest_range   = round(ranges[min_idx], 4)
            closest_angle   = round(
                math.degrees(msg.angle_min + min_idx * msg.angle_increment), 2
            )
        else:
            stat_min = stat_max = stat_mean = stat_std = None
            closest_range = closest_angle = None

        # ── Ghi Sheet 1 (1 hàng / scan) ───────────────────────────────────
        self.ws_summary.append([
            idx,
            ts,
            round(math.degrees(msg.angle_min), 3),
            round(math.degrees(msg.angle_max), 3),
            round(math.degrees(msg.angle_increment), 5),
            round(msg.range_min, 4),
            round(msg.range_max, 4),
            n,
            stat_min,
            stat_max,
            stat_mean,
            stat_std,
            closest_range,
            closest_angle,
        ])

        # ── Ghi Sheet 2 (1 hàng / beam) ───────────────────────────────────
        angle_min = msg.angle_min
        angle_inc = msg.angle_increment
        ws = self.ws_ranges
        for i, r in enumerate(ranges):
            angle_deg = round(math.degrees(angle_min + i * angle_inc), 3)
            range_val = round(r, 4) if math.isfinite(r) else None
            ws.append([idx, ts, i, angle_deg, range_val])

        self.get_logger().info(
            f"[{idx}] ts={ts:.3f}s | beams={n} | "
            f"min={stat_min}m | max={stat_max}m | mean={stat_mean}m"
        )

        if self.max_messages > 0 and self.message_count >= self.max_messages:
            self.get_logger().info(f"Đạt giới hạn {self.max_messages} messages → lưu file...")
            self._save()
            raise SystemExit

    # ─── Lưu file ─────────────────────────────────────────────────────────────
    def _save(self):
        try:
            # Freeze panes
            self.ws_summary.freeze_panes = "A2"
            self.ws_ranges.freeze_panes  = "A2"
            self.wb.save(self.output_path)
            self.get_logger().info(f"✓ Đã lưu: {self.output_path}")
        except Exception as e:
            self.get_logger().error(f"Lỗi khi lưu: {e}")


# ─── Main ─────────────────────────────────────────────────────────────────────
def main(args=None):
    rclpy.init(args=args)
    node = LidarToExcel()
    try:
        rclpy.spin(node)
    except (KeyboardInterrupt, SystemExit):
        pass
    finally:
        if node.message_count > 0:
            node._save()
        else:
            node.get_logger().warn("Không nhận được message nào.")
        node.destroy_node()
        rclpy.shutdown()


if __name__ == "__main__":
    main()
