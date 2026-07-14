#!/usr/bin/env python3
"""
depth_lidar_capture.py  –  Burst mode
──────────────────────────────────────
Nhấn Enter → nhập nhãn → thu liên tục 3 giây (≈90 frame @ 30fps):
  • Lưu ảnh màu  : photos/color_B001_F001_1m.png
  • Lưu ảnh depth: photos/depth_B001_F001_1m.png  (grayscale, gần = sáng)
  • Ghi Excel     : 1 hàng / frame

Excel – sheet "Measurements":
  burst | frame | label | timestamp | depth_center_m | depth_roi_mean_m |
  depth_roi_std_m | lidar_mean_m | lidar_min_m | lidar_count |
  color_file | depth_file

Yêu cầu: pip install openpyxl numpy
"""

import math, os, queue, struct, threading, time, zlib
import rclpy
from rclpy.node import Node
from sensor_msgs.msg import Image, LaserScan

try:
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise ImportError(f"pip install openpyxl numpy  ({e})")

# ── style helpers ─────────────────────────────────────────────────────────────
_FILL_BLUE  = PatternFill("solid", fgColor="1F4E79")
_FONT_HDR   = Font(bold=True, color="FFFFFF")
_CENTER     = Alignment(horizontal="center")

def _style_header(ws, fill):
    for c in ws[1]:
        c.font = _FONT_HDR; c.fill = fill; c.alignment = _CENTER

# ── image helpers (pure Python, no cv2) ──────────────────────────────────────
def _decode_depth(msg: Image) -> np.ndarray:
    if msg.encoding == "32FC1":
        return np.frombuffer(bytes(msg.data), np.float32).reshape(msg.height, msg.width).copy()
    elif msg.encoding in ("16UC1", "mono16"):
        return (np.frombuffer(bytes(msg.data), np.uint16)
                  .reshape(msg.height, msg.width).astype(np.float32) / 1000.0)
    raise RuntimeError(f"Encoding không hỗ trợ: {msg.encoding}")

def _to_png(arr_uint8_rgb: np.ndarray) -> bytes:
    """Chuyển H×W×3 uint8 RGB → PNG bytes (pure Python)."""
    H, W = arr_uint8_rgb.shape[:2]
    def chunk(t, d):
        c = t + d
        return struct.pack(">I", len(d)) + c + struct.pack(">I", zlib.crc32(c) & 0xFFFFFFFF)
    ihdr = chunk(b"IHDR", struct.pack(">IIBBBBB", W, H, 8, 2, 0, 0, 0))
    raw  = b"".join(b"\x00" + arr_uint8_rgb[r].tobytes() for r in range(H))
    idat = chunk(b"IDAT", zlib.compress(raw, 1))   # level=1: nhanh hơn
    iend = chunk(b"IEND", b"")
    return b"\x89PNG\r\n\x1a\n" + ihdr + idat + iend

def _color_msg_to_png(msg: Image) -> bytes:
    arr = np.frombuffer(bytes(msg.data), np.uint8).reshape(msg.height, msg.width, 3)
    if msg.encoding == "bgr8":
        arr = arr[:, :, ::-1].copy()
    return _to_png(arr)

def _depth_arr_to_png(arr: np.ndarray, d_min=0.1, d_max=10.0) -> bytes:
    """Depth float32 (m) → grayscale PNG: gần = sáng."""
    valid = np.isfinite(arr) & (arr > 0)
    g = np.zeros_like(arr, dtype=np.uint8)
    g[valid] = np.clip(
        255 * (1.0 - (arr[valid] - d_min) / (d_max - d_min)), 0, 255
    ).astype(np.uint8)
    rgb = np.stack([g, g, g], axis=-1)
    return _to_png(rgb)


class DepthLidarCapture(Node):
    def __init__(self):
        super().__init__("depth_lidar_capture_node")

        # ── params ────────────────────────────────────────────────────────
        def P(n, v): self.declare_parameter(n, v); return self.get_parameter(n).value

        self.output_path     = P("output_path", os.path.expanduser("~/depth_lidar_test.xlsx"))
        self.photo_dir       = P("photo_dir",   os.path.expanduser("~/depth_lidar_photos"))
        self.depth_topic     = P("depth_topic", "/camera/camera/aligned_depth_to_color/image_raw")
        self.color_topic     = P("color_topic", "/camera/camera/color/image_raw")
        self.roi_radius      = int(P("roi_radius_px",   20))
        self.front_angle_deg = float(P("front_angle_deg", 10.0))
        self.burst_sec       = float(P("burst_sec",       3.0))
        self.depth_vis_max   = float(P("depth_vis_max_m", 10.0))

        os.makedirs(self.photo_dir, exist_ok=True)

        self.get_logger().info(f"Output  : {self.output_path}")
        self.get_logger().info(f"Photos  : {self.photo_dir}")
        self.get_logger().info(f"Burst   : {self.burst_sec}s mỗi lần Enter")

        # ── state ─────────────────────────────────────────────────────────
        self.latest_depth: np.ndarray = None
        self.latest_color: Image      = None
        self.latest_scan:  LaserScan  = None
        self._dlck = threading.Lock()
        self._clck = threading.Lock()
        self._slck = threading.Lock()

        self.burst_active   = False
        self.burst_end      = 0.0
        self.burst_label    = ""
        self.burst_id       = 0
        self.burst_frame    = 0
        self._blck          = threading.Lock()

        self.total_rows     = 0

        # ── workbook ──────────────────────────────────────────────────────
        self.wb   = Workbook()
        self.ws   = self.wb.active
        self.ws.title = "Measurements"
        self._init_header()

        # ── save queue (background writer) ────────────────────────────────
        self._q = queue.Queue()
        self._writer = threading.Thread(target=self._save_worker, daemon=True)
        self._writer.start()

        # ── subscribers ───────────────────────────────────────────────────
        self.create_subscription(Image,     self.depth_topic, self._depth_cb, 10)
        self.create_subscription(Image,     self.color_topic, self._color_cb, 10)
        self.create_subscription(LaserScan, "/scan",          self._scan_cb,  10)

        # ── keyboard ──────────────────────────────────────────────────────
        threading.Thread(target=self._kb_loop, daemon=True).start()

        print("\n" + "="*60)
        print("  BURST DEPTH + COLOR + LIDAR CAPTURE")
        print("="*60)
        print(f"  Nhấn [Enter] → nhập nhãn → thu {self.burst_sec}s liên tục")
        print("  Nhấn [Ctrl+C] để lưu và thoát")
        print("="*60 + "\n")

    # ── callbacks ────────────────────────────────────────────────────────────
    def _depth_cb(self, msg: Image):
        try:
            arr = _decode_depth(msg)
        except Exception as e:
            self.get_logger().error(f"Depth decode: {e}"); return

        with self._dlck:
            self.latest_depth = arr

        # trigger burst frame
        with self._blck:
            active = self.burst_active and (time.monotonic() < self.burst_end)

        if active:
            with self._clck: color_snap = self.latest_color
            with self._slck: scan_snap  = self.latest_scan
            if scan_snap is not None:
                self._queue_frame(arr, color_snap, scan_snap)

        elif self.burst_active and time.monotonic() >= self.burst_end:
            with self._blck:
                if self.burst_active:   # chỉ print 1 lần
                    self.burst_active = False
                    n = self.burst_frame
                    b = self.burst_id
            print(f"\n  ✓ Burst #{b} kết thúc: {n} frames đã ghi vào queue.")
            print("  → Nhấn Enter để đo đợt tiếp ...\n")

    def _color_cb(self, msg: Image):
        with self._clck: self.latest_color = msg

    def _scan_cb(self, msg: LaserScan):
        with self._slck: self.latest_scan = msg

    # ── burst frame ──────────────────────────────────────────────────────────
    def _queue_frame(self, depth_arr, color_msg, scan_msg):
        with self._blck:
            self.burst_frame += 1
            burst_id    = self.burst_id
            frame_idx   = self.burst_frame
            label       = self.burst_label

        stamp = scan_msg.header.stamp
        ts    = round(stamp.sec + stamp.nanosec * 1e-9, 6)

        # depth stats
        H, W = depth_arr.shape[:2]
        cx, cy, r = W//2, H//2, self.roi_radius
        roi = depth_arr[max(0,cy-r):cy+r+1, max(0,cx-r):cx+r+1]
        mask = np.isfinite(roi) & (roi > 0.01)
        vals = roi[mask]
        if vals.size:
            cv = roi[min(r,cy), min(r,cx)]
            d_center = round(float(cv),   4) if np.isfinite(cv) and cv>0 else None
            d_mean   = round(float(np.mean(vals)), 4)
            d_std    = round(float(np.std(vals)),  4)
        else:
            d_center = d_mean = d_std = None

        # lidar front
        fa = math.radians(self.front_angle_deg)
        lr = [v for i,v in enumerate(scan_msg.ranges)
              if math.isfinite(v) and v>0
              and abs(scan_msg.angle_min + i*scan_msg.angle_increment) <= fa]
        if lr:
            l_mean = round(float(np.mean(lr)), 4)
            l_min  = round(float(min(lr)),     4)
            l_n    = len(lr)
        else:
            l_mean = l_min = l_n = None

        cfn   = f"color_B{burst_id:03d}_F{frame_idx:03d}.png"
        dfn   = f"depth_B{burst_id:03d}_F{frame_idx:03d}.png"

        # enqueue (encode PNG in background)
        self._q.put({
            "row":   [burst_id, frame_idx, ts,
                      d_center, d_mean, d_std,
                      l_mean, l_min, l_n, cfn, dfn],
            "depth_arr":  depth_arr,
            "color_msg":  color_msg,
            "color_file": os.path.join(self.photo_dir, cfn),
            "depth_file": os.path.join(self.photo_dir, dfn),
        })

    # ── background writer ────────────────────────────────────────────────────
    def _save_worker(self):
        while True:
            item = self._q.get()
            if item is None: break
            try:
                # color PNG
                if item["color_msg"] is not None:
                    with open(item["color_file"], "wb") as f:
                        f.write(_color_msg_to_png(item["color_msg"]))
                # depth PNG
                with open(item["depth_file"], "wb") as f:
                    f.write(_depth_arr_to_png(item["depth_arr"],
                                              d_max=self.depth_vis_max))
                # excel row
                self.ws.append(item["row"])
                self.total_rows += 1
            except Exception as e:
                self.get_logger().error(f"Save worker: {e}")
            self._q.task_done()

    # ── keyboard ─────────────────────────────────────────────────────────────
    def _kb_loop(self):
        while rclpy.ok():
            try:
                input("  → Nhấn Enter để bắt đầu burst ... ")
            except EOFError:
                break

            with self._dlck: have_d = self.latest_depth is not None
            with self._slck: have_s = self.latest_scan  is not None
            if not have_d or not have_s:
                print("  ⚠  Chưa có dữ liệu depth/LiDAR. Thử lại sau.\n")
                continue

            with self._blck:
                self.burst_id    += 1
                self.burst_frame  = 0
                self.burst_end    = time.monotonic() + self.burst_sec
                self.burst_active = True
                bid = self.burst_id

            print(f"  ▶ Burst #{bid} – thu trong {self.burst_sec}s ...\n")

    # ── header ───────────────────────────────────────────────────────────────
    def _init_header(self):
        cols = ["burst", "frame", "timestamp_sec",
                "depth_center_m", "depth_roi_mean_m", "depth_roi_std_m",
                "lidar_mean_m", "lidar_min_m", "lidar_count",
                "color_file", "depth_file"]
        self.ws.append(cols)
        _style_header(self.ws, _FILL_BLUE)
        widths = [8, 8, 18, 16, 18, 16, 14, 12, 12, 32, 32]
        for i,w in enumerate(widths,1):
            self.ws.column_dimensions[get_column_letter(i)].width = w
        self.ws.freeze_panes = "A2"

    # ── save ─────────────────────────────────────────────────────────────────
    def save(self):
        self._q.join()          # đợi queue flush hết
        self._q.put(None)       # stop worker
        try:
            self.wb.save(self.output_path)
            print(f"\n  ✓ Excel ({self.total_rows} hàng): {self.output_path}")
            print(f"  ✓ Ảnh              : {self.photo_dir}/")
        except Exception as e:
            self.get_logger().error(f"Lưu Excel: {e}")


# ── main ─────────────────────────────────────────────────────────────────────
def main(args=None):
    rclpy.init(args=args)
    node = DepthLidarCapture()
    try:
        rclpy.spin(node)
    except KeyboardInterrupt:
        pass
    finally:
        node.save()
        node.destroy_node()
        if rclpy.ok():
            rclpy.shutdown()

  
if __name__ == "__main__":
    main()
